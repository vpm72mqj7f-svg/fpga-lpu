"""
gen_sizing_excel.py -- Generate traffic-driven super-node sizing Excel spreadsheet.
Output: docs/fpga_supernode_sizing.xlsx
"""

import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from copy import copy


# ============================================================================
# Constants
# ============================================================================
CHIPS_PER_SERVER   = 32
HBM_GB_PER_CHIP    = 32
WEIGHT_GB_PER_CHIP = 0.7         # fp4 expert+attn+router per chip
HBM_KV_AVAIL_GB    = HBM_GB_PER_CHIP - WEIGHT_GB_PER_CHIP  # 31.3
NUM_LAYERS         = 61
MLA_KV_BYTES       = 576          # bytes per token per layer
V4_ACTIVE_SCALE    = 49 / 37      # V4 Pro 49B active vs V3 37B
PIPELINE_TPS       = int(23_104 / V4_ACTIVE_SCALE)  # ~17,445
BATCH1_TPS         = int(875 / V4_ACTIVE_SCALE)      # ~660
TOKEN_LATENCY_US   = int(1140 * V4_ACTIVE_SCALE)     # ~1,510

# FP8 算力归一化 (与 GPU 对比)
DSP_FP8_TFLOPS_PER_CHIP = 12_300 * 450 * 1 / 1e6 * 2   # 11.07 TFLOPS (fp8×fp8 保守)
FPGA_FP8_TFLOPS_PER_SRV = DSP_FP8_TFLOPS_PER_CHIP * CHIPS_PER_SERVER  # ~354 TFLOPS
H200_FP8_TFLOPS         = 1_979     # H200 SXM, FP8 dense
H200_FP8_TFLOPS_PER_SRV = 1_979 * 8 # 15,832 TFLOPS
ASCEND_FP8_TFLOPS       = 800       # Ascend 950PR, FP8 (估算)
ASCEND_FP8_TFLOPS_PER_SRV = 800 * 8 # 6,400 TFLOPS
# 带宽/算力比 (Memory Bandwidth per Compute, GB/s per TFLOP)
FPGA_MBW  = 920   / DSP_FP8_TFLOPS_PER_CHIP   # ~83
H200_MBW  = 4_800 / H200_FP8_TFLOPS           # ~2.4
ASCEND_MBW = 4_000 / ASCEND_FP8_TFLOPS        # ~5.0

SERVER_COST_RMB    = 1_000_000  # ~57.6万 chips + 42.4万 BOM/chassis
A7_CHIP_COST_USD   = 2_500      # AGM 039-F per-chip, USD (excl. tax)
A7_CHIP_COST_RMB   = 18_000     # ~$2,500 x 7.2 RMB/USD
A7_CHIPS_PER_SRV   = 32
A7_CHIP_TOTAL_RMB  = A7_CHIP_COST_RMB * A7_CHIPS_PER_SRV  # 576,000
SERVER_POWER_KW    = 5.3
RMB_PER_KWH        = 0.35     # 东数西算/内蒙数据中心


def sessions_per_server(context_len, target_tps=30):
    """Max sessions per server at given context length and per-user TPS target."""
    # Memory bound: per-chip KV cache for 2-layer chips
    kv_per_chip_2layer_gb = 2 * MLA_KV_BYTES * context_len / (1024**3)
    sessions_mem = int(HBM_KV_AVAIL_GB / kv_per_chip_2layer_gb)
    # Compute bound
    sessions_cmp = int(PIPELINE_TPS / target_tps)
    return min(sessions_mem, sessions_cmp), "memory" if sessions_mem <= sessions_cmp else "compute"


# ============================================================================
# Styles
# ============================================================================
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TITLE_FONT = Font(name="Microsoft YaHei", bold=True, size=14, color="1F3864")
SECTION_FONT = Font(name="Microsoft YaHei", bold=True, size=12, color="2F5496")
BODY_FONT = Font(name="Microsoft YaHei", size=10)
BOLD_FONT = Font(name="Microsoft YaHei", size=10, bold=True)
NUM_FONT = Font(name="Consolas", size=10)
RESULT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
BAD_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)


def style_header_row(ws, row, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = thin_border


def style_data_cell(ws, row, col, fmt=None, font=None, fill=None):
    c = ws.cell(row=row, column=col)
    c.font = font or BODY_FONT
    c.border = thin_border
    c.alignment = CENTER
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================================
# Sheet 1: Traffic-Driven Sizing Calculator
# ============================================================================

def build_calculator_sheet(wb):
    ws = wb.active
    ws.title = "流量驱动规模计算器"

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "FPGA 超节点规模计算器 — 按业务流量定义服务器数量"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # -- Input Section --
    ws.merge_cells('A3:H3')
    ws['A3'] = "输入参数 (黄色单元格可修改)"
    ws['A3'].font = SECTION_FONT

    inputs = [
        ("B5", "并发会话数", 1000, "#,##0", "同时在线进行推理的会话数"),
        ("B6", "上下文长度 (tokens)", 262144, "#,##0", "每会话 KV cache 窗口大小 (标配256K)"),
        ("B7", "每用户目标 tok/s", 30, "0", "保证每个会话的最低输出速率"),
        ("B8", "突发系数", 1.3, "0.0", "峰值/均值比,用于容量缓冲"),
        ("B9", "服务器单价 (RMB)", 1200000, "#,##0", "每台 4U FPGA 服务器成本"),
        ("B10", "每台功耗 (kW)", 5.3, "0.0", "含 8 卡 x 4 芯片 + CPU + 散热"),
        ("B11", "电费 (RMB/kWh)", 0.8, "0.00", "数据中心电价"),
        ("B12", "PUE", 1.3, "0.0", "电源使用效率"),
    ]

    for cell, label, default, fmt, note in inputs:
        r = int(cell[1:])
        ws[f'A{r}'] = label
        ws[f'A{r}'].font = BOLD_FONT
        ws[f'A{r}'].alignment = LEFT
        ws[f'A{r}'].border = thin_border

        ws[cell] = default
        style_data_cell(ws, r, 2, fmt=fmt, fill=INPUT_FILL)

        ws.cell(row=r, column=3, value=note)
        ws.cell(row=r, column=3).font = BODY_FONT
        ws.cell(row=r, column=3).alignment = LEFT

    set_col_widths(ws, [22, 22, 40, 18, 18, 18, 18, 18])

    # -- Calculation Section --
    ws.merge_cells('A15:H15')
    ws['A15'] = "中间计算"
    ws['A15'].font = SECTION_FONT

    calc_rows = [
        ("B16", "每 chip KV cache 可用 (GB)",   f"={HBM_KV_AVAIL_GB}"),
        ("B17", "每 token 每层 KV 字节",         f"={MLA_KV_BYTES}"),
        ("B18", "每会话 KV cache (GB)",          "=B6*B17*61/1024^3"),
        ("B19", "每 chip 2 层 KV/会话 (GB)",      "=B6*B17*2/1024^3"),
        ("B20", "显存约束: 每 chip 会话数",        "=INT(B16/B19)"),
        ("B21", "算力约束: 每台会话数",            f"=INT({PIPELINE_TPS}/B7)"),
        ("B22", "每台最大会话数",                  "=MIN(B20,B21)"),
        ("B23", "单台吞吐 (tok/s)",               f"={PIPELINE_TPS}"),
    ]

    for cell, label, formula in calc_rows:
        r = int(cell[1:])
        ws[f'A{r}'] = label
        ws[f'A{r}'].font = BODY_FONT
        ws[f'A{r}'].alignment = LEFT
        ws[f'A{r}'].border = thin_border

        ws[cell] = formula if formula.startswith('=') else float(formula)
        style_data_cell(ws, r, 2, fmt="#,##0.00" if "GB" in label else "#,##0")

    # -- Result Section --
    ws.merge_cells('A26:H26')
    ws['A26'] = "计算结果"
    ws['A26'].font = SECTION_FONT

    result_rows = [
        ("A28", "会话并发需要 (台)",    "=ROUNDUP(B5/B22,0)"),
        ("A29", "吞吐需要 (台)",        "=ROUNDUP(B5*B7/B23,0)"),
        ("A30", "突发需要 (台)",        "=ROUNDUP(B5*B8/B22,0)"),
        ("A31", "=> 超节点服务器总数",  "=MAX(B28,B29,B30)", True),
        ("A32", "=> 所需机柜数 (9台/柜)","=ROUNDUP(B31/9,0)", True),
        ("A34", "总吞吐 (tok/s)",       "=B31*B23"),
        ("A35", "每用户实际 tok/s",     "=B34/B5"),
        ("A36", "DSP 利用率",           "=B5*B7/B34"),
        ("A37", "富余会话数",           "=B31*B22-B5"),
        ("A38", "KV Cache 总量 (GB)",   "=B18*B5"),
        ("A40", "服务器 CAPEX (RMB)",   "=B31*B9"),
        ("A41", "年电费 (RMB)",         "=B31*B10*8760*B11*B12"),
        ("A42", "年运维 (RMB)",         "=B40*0.05"),
        ("A43", "年 OPEX (RMB)",        "=B41+B42"),
        ("A44", "5 年 TCO (RMB)",       "=B40+B43*5"),
        ("A45", "每会话/年成本 (RMB)",  "=B44/5/B5"),
        ("A46", "每百万 token 成本 (RMB)","=B45/365/86400/B7*1e6"),
    ]

    for cell, label, formula, *rest in result_rows:
        is_key = rest[0] if rest else False
        r = int(cell[1:])
        ws[f'A{r}'] = label
        ws[f'A{r}'].font = BOLD_FONT if is_key else BODY_FONT
        ws[f'A{r}'].alignment = LEFT
        ws[f'A{r}'].border = thin_border

        ws.cell(row=r, column=2, value=formula)
        style_data_cell(ws, r, 2, fmt="#,##0",
                        fill=RESULT_FILL if is_key else None,
                        font=BOLD_FONT if is_key else NUM_FONT)

    # -- Context Length Sensitivity --
    ws.merge_cells('A50:H50')
    ws['A50'] = "上下文长度灵敏度分析"
    ws['A50'].font = SECTION_FONT

    ctx_headers = ["上下文长度", "会话/台", "约束", "500并发需几台", "1000并发", "5000并发", "10000并发", "50000并发"]
    for i, h in enumerate(ctx_headers, 1):
        ws.cell(row=52, column=i, value=h)
    style_header_row(ws, 52, len(ctx_headers))

    ctx_list = [4096, 8192, 16384, 32768, 65536, 131072, 262144]
    target_tps = 30
    for idx, ctx in enumerate(ctx_list):
        row = 53 + idx
        sps, bound = sessions_per_server(ctx, target_tps)
        ws.cell(row=row, column=1, value=f"{ctx//1024}K")
        style_data_cell(ws, row, 1)

        ws.cell(row=row, column=2, value=sps)
        style_data_cell(ws, row, 2, fmt="#,##0")

        ws.cell(row=row, column=3, value={"memory":"显存", "compute":"算力"}[bound])
        style_data_cell(ws, row, 3)

        for j, sessions in enumerate([500, 1000, 5000, 10000, 50000]):
            n = math.ceil(sessions / sps)
            ws.cell(row=row, column=4+j, value=n)
            style_data_cell(ws, row, 4+j, fmt="#,##0")

    # Data bars on server counts
    ws.conditional_formatting.add(
        'D53:H59',
        DataBarRule(start_type='min', end_type='max', color='2F5496', showValue=True)
    )

    # Protect only formula cells conceptually (just color them)
    for r in range(1, 60):
        ws.row_dimensions[r].height = 20


# ============================================================================
# Sheet 2: Scenario Comparison
# ============================================================================

def build_scenarios_sheet(wb):
    ws = wb.create_sheet("场景对比")

    ws.merge_cells('A1:K1')
    ws['A1'] = "典型业务场景对比 — 按流量定义超节点规模"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    scenarios = [
        ("PoC 验证", 200, 262144, 30, 1.3, 1024, 512),
        ("小规模云", 1000, 262144, 30, 1.3, 1024, 512),
        ("中规模云", 5000, 262144, 30, 1.3, 1024, 512),
        ("大规模云", 20000, 262144, 30, 1.3, 1024, 512),
        ("DeepSeek 公开服务量级", 100000, 131072, 25, 1.5, 512, 256),
        ("企业专属 Premium", 500, 262144, 50, 1.2, 4096, 2048),
        ("轻量 chat (短上下文)", 3000, 32768, 30, 1.3, 512, 256),
        ("代码助手 (长上下文)", 1000, 262144, 35, 1.2, 2048, 1024),
    ]

    headers = [
        "场景", "并发会话", "上下文(K)", "目标tok/s", "突发系数",
        "会话/台", "约束", "需服务器", "需机柜",
        "总吞吐(tok/s)", "每用户TPS", "DSP利用率",
        "5年TCO(万)", "每会话/年(元)", "每百万token(元)"
    ]

    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(headers))

    for idx, (name, sessions, ctx, tps, burst, prompt, output) in enumerate(scenarios):
        row = 4 + idx
        sps, bound = sessions_per_server(ctx, tps)
        servers = max(
            math.ceil(sessions / sps),
            math.ceil(sessions * tps / PIPELINE_TPS),
            math.ceil(sessions * burst / sps),
        )
        racks = math.ceil(servers / 9)
        agg_tps = servers * PIPELINE_TPS
        per_user_tps = agg_tps / sessions if sessions > 0 else 0
        dsp_util = sessions * tps / agg_tps
        capex = servers * SERVER_COST_RMB / 1e4
        annual_opex = servers * SERVER_POWER_KW * 8760 * RMB_PER_KWH * 1.3 / 1e4
        tco_5yr = capex + annual_opex * 5
        cost_per_session_yr = tco_5yr * 1e4 / 5 / sessions if sessions > 0 else 0
        cost_per_1m_tokens = cost_per_session_yr / 365 / 86400 / tps * 1e6 if tps > 0 else 0

        vals = [
            name, sessions, f"{ctx//1024}K", tps, f"{burst}x",
            sps, {"memory":"显存","compute":"算力"}[bound], servers, racks,
            f"{agg_tps:,.0f}", f"{per_user_tps:.0f}", f"{dsp_util:.1%}",
            f"{tco_5yr:.0f}", f"{cost_per_session_yr:.0f}", f"{cost_per_1m_tokens:.2f}",
        ]

        for j, v in enumerate(vals, 1):
            ws.cell(row=row, column=j, value=v)
            c = ws.cell(row=row, column=j)
            c.font = BODY_FONT
            c.border = thin_border
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if j == 8:  # server count
                c.font = BOLD_FONT
                c.fill = RESULT_FILL

    # Column widths
    widths = [22, 12, 12, 12, 10, 12, 8, 10, 8, 16, 12, 12, 14, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for r in range(1, 15):
        ws.row_dimensions[r].height = 22

    # -- TCO comparison bar (manual) --
    r = 14
    ws.merge_cells(f'A{r}:P{r}')
    ws[f'A{r}'] = "注: 服务器单价 120 万 RMB/台, 电价 0.8 元/kWh, PUE=1.3, 运维=CAPEX*5%/年. 超节点规模由会话并发+吞吐+突发三者取最大值."
    ws[f'A{r}'].font = Font(name="Microsoft YaHei", size=9, color="666666")

    # Add scenario description column for reference
    ws.merge_cells(f'A16:P16')
    ws[f'A16'] = "场景说明: PoC=概念验证(1-2台即可) | 小规模云=初创公司/内部工具(几台) | 中规模=对外SaaS服务(数十台) | 大规模=头部云厂商(百台级) | DeepSeek量级=千万DAU(百台+)"
    ws[f'A16'].font = Font(name="Microsoft YaHei", size=9, color="666666")


# ============================================================================
# Sheet 3: Batch Throughput Model
# ============================================================================

def build_batch_sheet(wb):
    ws = wb.create_sheet("Continuous Batching")

    ws.merge_cells('A1:G1')
    ws['A1'] = "Continuous Batching 吞吐模型"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:G2')
    K = PIPELINE_TPS / BATCH1_TPS - 1
    ws['A2'] = f"Pipeline 效率 = B/(B+K), K={K:.1f}. 瓶颈 chip 时间 = 37.4 us (2层 x 18.7 us)."
    ws['A2'].font = Font(name="Microsoft YaHei", size=9, color="666666")

    headers = ["Batch Size B", "Pipeline 效率", "总 TPS", "TPOT (ms)", "每请求 TPS",
               "DSP 利用率", "优于 30 tok/s?"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, len(headers))

    batch_sizes = [1, 2, 4, 8, 16, 32, 64, 128, 256]
    for idx, B in enumerate(batch_sizes):
        row = 5 + idx
        eff = B / (B + K)
        tps = PIPELINE_TPS * eff
        tpot = 1000.0 * B / tps if tps > 0 else 0
        per_req = tps / B
        dsp = 0.589 * eff

        ws.cell(row=row, column=1, value=B)
        style_data_cell(ws, row, 1)

        ws.cell(row=row, column=2, value=eff)
        style_data_cell(ws, row, 2, fmt="0.0%")

        ws.cell(row=row, column=3, value=tps)
        style_data_cell(ws, row, 3, fmt="#,##0")

        ws.cell(row=row, column=4, value=tpot)
        style_data_cell(ws, row, 4, fmt="0.00")

        ws.cell(row=row, column=5, value=per_req)
        style_data_cell(ws, row, 5, fmt="#,##0")

        ws.cell(row=row, column=6, value=dsp)
        style_data_cell(ws, row, 6, fmt="0.0%")

        ok = "YES" if per_req >= 30 else "NO"
        ws.cell(row=row, column=7, value=ok)
        style_data_cell(ws, row, 7,
                        fill=RESULT_FILL if ok == "YES" else WARN_FILL,
                        font=BOLD_FONT)

    set_col_widths(ws, [16, 16, 14, 12, 14, 14, 18])

    # Prefill latency table
    ws.merge_cells('A18:G18')
    ws['A18'] = "Prefill 延迟 (prefill_tps ≈ 20,300 tok/s, pipeline fill ≈ 1.2ms)"
    ws['A18'].font = SECTION_FONT

    pf_headers = ["Prompt tokens", "Prefill (ms)", "+ B=16 TPOT", "TTFT B=16",
                  "+ B=64 TPOT", "TTFT B=64", "+ B=256 TPOT", "TTFT B=256"]
    for i, h in enumerate(pf_headers, 1):
        ws.cell(row=20, column=i, value=h)
    style_header_row(ws, 20, len(pf_headers))

    prefill_tps = PIPELINE_TPS * 0.88
    fill_ms = 31 * 2 * 18.7 / 1000  # ~1.16ms
    tpot_16 = 1000.0 * 16 / (PIPELINE_TPS * 16/(16+K))
    tpot_64 = 1000.0 * 64 / (PIPELINE_TPS * 64/(64+K))
    tpot_256 = 1000.0 * 256 / (PIPELINE_TPS * 256/(256+K))

    for idx, pl in enumerate([256, 512, 1024, 2048, 4096, 8192, 16384, 32768]):
        row = 21 + idx
        prefill_ms = pl / prefill_tps * 1000 + fill_ms
        vals = [pl, prefill_ms, tpot_16, prefill_ms+tpot_16, tpot_64, prefill_ms+tpot_64, tpot_256, prefill_ms+tpot_256]
        for j, v in enumerate(vals, 1):
            ws.cell(row=row, column=j, value=v)
            style_data_cell(ws, row, j, fmt="0.0" if j > 1 else "#,##0")

    set_col_widths(ws, [16, 14, 14, 14, 14, 14, 14, 14])


# ============================================================================
# Sheet 4: GPU vs FPGA
# ============================================================================

def build_comparison_sheet(wb):
    ws = wb.create_sheet("FPGA vs GPU vs NPU")

    ws.merge_cells('A1:E1')
    ws['A1'] = "FPGA 超节点 vs GPU vs NPU — 三路云服务对比"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    headers = ["指标", "FPGA 超节点 (A7 x1024)", "H100 x256", "Ascend 950PR x256 (估)", "FPGA 优势"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, len(headers))

    # Note: Ascend 950PR specs are estimates based on Ascend 910B/C evolution
    # FP16 ~400 TFLOPS, FP8 ~800 TFLOPS, HBM3 128GB, ~4TB/s, ~450W
    # Per-card price estimate: ~120,000 RMB (constrained supply)

    rows = [
        ("服务器形态", "4U, 8卡x4芯片=32 FPGA", "4U, 8 GPU", "4U, 8 NPU", "密度最高"),
        ("芯片总数", "1,024 FPGA", "256 H100", "256 Ascend 950PR", "芯片数 4x"),
        ("单芯片价格", f"$2,500 (~RMB {A7_CHIP_COST_RMB/1000:.0f}K)", "~$30K (~RMB 21万)", "~RMB 12万 (估)", "8.4x 更便宜 vs H100"),
        ("单芯片功耗", "~166 W", "~700 W", "~450 W (估)", "2.7x 更低 vs 950PR"),
        ("单芯片 HBM", "32 GB", "141 GB", "128 GB (估)", "—"),
        ("HBM 带宽", "920 GB/s", "4.8 TB/s", "~4 TB/s (估)", "—"),
        ("单芯片 FP8 TFLOPS (归一化)", f"{DSP_FP8_TFLOPS_PER_CHIP:.1f} (fp4xfp8→fp8×fp8)", f"{H200_FP8_TFLOPS:,} (原生 fp8)", f"{ASCEND_FP8_TFLOPS:,} (估算)", f"H200 算力 {H200_FP8_TFLOPS/DSP_FP8_TFLOPS_PER_CHIP:.0f}x"),
        ("单服务器 FP8 TFLOPS", f"{FPGA_FP8_TFLOPS_PER_SRV:.0f}", f"{H200_FP8_TFLOPS_PER_SRV:,}", f"{ASCEND_FP8_TFLOPS_PER_SRV:,}", f"H200 算力 {H200_FP8_TFLOPS_PER_SRV/FPGA_FP8_TFLOPS_PER_SRV:.0f}x"),
        ("带宽/算力比 (GB/s per TFLOP)", f"{FPGA_MBW:.1f} (memory-bound 友好)", f"{H200_MBW:.1f} (compute 偏重)", f"{ASCEND_MBW:.1f} (compute 偏重)", f"FPGA 是 H200 的 {FPGA_MBW/H200_MBW:.0f}x"),
        ("单芯片推理 tok/s", f"{PIPELINE_TPS//CHIPS_PER_SERVER} (fp4xfp8, V4 scaled)", "~500 (fp8, TP=8)", "~312 (估, fp8)", "1.1-1.7x vs 竞品"),
        ("单服务器吞吐", f"{PIPELINE_TPS:,} tok/s (fp4xfp8, V4 Pro)", "~4,000 tok/s", "~2,500 tok/s (估)", "4.4-7.0x"),
        ("32台总吞吐", f"{PIPELINE_TPS*32/1000:.0f}K tok/s", "~128K tok/s", "~80K tok/s (估)", "4.4-7.0x"),
        ("单机会话数 @256K", f"{sessions_per_server(262144, 30)[0]}", "~40 (权重+KV争HBM)", "~30 (估)", f"{sessions_per_server(262144, 30)[0]/15:.1f}x vs H200"),
        ("每 token 延迟", f"{TOKEN_LATENCY_US/1000:.2f} ms", "~8 ms (TP=8)", "~12 ms (估, TP=8)", "5.3x 更低 vs H200"),
        ("互联需求", "无 (数据并行)", "InfiniBand NDR400", "HCCS + RoCE", "架构大幅简化"),
        ("互联成本/服务器", "0", "~30万 RMB", "~20万 RMB (估)", "省数百万元"),
        ("整机价格 (含交换机)", f"RMB {SERVER_COST_RMB/1e4:.0f}万", "~RMB 300万", "~RMB 130万 (估)", "1.3-3x 更便宜"),
        ("每 tok/s 系统成本", f"~{SERVER_COST_RMB/PIPELINE_TPS:.0f} RMB", "~825 RMB", "~600 RMB (估)", "14-19x 更低"),
        ("模型加载方式", "单台完整模型 (fp4)", "需 TP=4 分片", "需 TP=4 分片 (估)", "架构简化"),
        ("集群目的", "提升并发吞吐", "装下模型 + 吞吐", "装下模型 + 吞吐", "根本差异"),
        ("供应风险", "国产 AGM, 可控", "美国出口管制", "国产, 供应紧张", "供应链安全"),
    ]

    for idx, (metric, fpga, gpu, ascend, advantage) in enumerate(rows):
        row = 4 + idx
        ws.cell(row=row, column=1, value=metric)
        style_data_cell(ws, row, 1, font=BOLD_FONT)
        ws.cell(row=row, column=1).alignment = LEFT

        ws.cell(row=row, column=2, value=fpga)
        style_data_cell(ws, row, 2)

        ws.cell(row=row, column=3, value=gpu)
        style_data_cell(ws, row, 3)

        ws.cell(row=row, column=4, value=ascend)
        style_data_cell(ws, row, 4)

        ws.cell(row=row, column=5, value=advantage)
        style_data_cell(ws, row, 5, fill=RESULT_FILL, font=BOLD_FONT)

    set_col_widths(ws, [24, 30, 28, 28, 20])
    for r in range(1, 27):
        ws.row_dimensions[r].height = 22

    # Key insight box
    ws.merge_cells('A27:E29')
    ws['A27'] = (
        "核心架构差异 (按重要性):\n"
        "1. 专家分布: FPGA 32 chip 分担 384 experts → 每 chip 仅 0.7 GB 专家权重。GPU TP=8 → 每 GPU 125 GB → KV 可用空间差 219x\n"
        "2. 算力归一化: FPGA DSP 11.07 TMACs (fp4×fp8) → 归一化 fp8×fp8 保守 5.54 TMACs → 11.07 TFLOPS (GPU 等效)。算力比 H200:FPGA = 45:1, 但带宽比仅 1.3:1 — LLM 推理是 memory-bound, 不是 compute-bound。FPGA 每 TFLOP 带宽 83 GB/s, H200 仅 2.4 GB/s\n"
        "3. 确定性权重 SRAM 缓存: Attention + Shared Expert + Router (~21 MB) 在 SRAM, 不走 HBM。GPU 这些也走 HBM\n"
        "4. fp4 专家权重: 33 MB/expert vs FP8 66 MB → 减半 (辅助因素, 非核心)\n"
        "5. 互联: FPGA 单机可装完整模型, 不需要 InfiniBand/HCCS 数据面\n"
        "A7 芯片 $2,500/片 x 32 ≈ 57.6 万芯片成本, 整机 ~100 万 RMB。H200 8-GPU ~300万, Ascend 8-NPU ~130万。"
    )
    ws['A27'].font = Font(name="Microsoft YaHei", size=10, color="1F3864")
    ws['A27'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[27].height = 80

    # Chip cost breakdown
    ws.merge_cells('A31:E31')
    ws['A31'] = "A7 FPGA 服务器成本拆解"
    ws['A31'].font = SECTION_FONT

    cost_headers = ["项目", "单价 (RMB)", "数量", "小计 (RMB)", "备注"]
    for i, h in enumerate(cost_headers, 1):
        ws.cell(row=33, column=i, value=h)
    style_header_row(ws, 33, len(cost_headers))

    cost_rows = [
        ("AGM 039-F (A7) 芯片", f"$2,500 (~{A7_CHIP_COST_RMB//1000}K RMB)", 32, f"{A7_CHIP_TOTAL_RMB:,d}", f"$2,500 x 32 = $80,000"),
        ("PCB 基板 (16层, 高速)", "15,000", 8, "120,000", "每 card 一片"),
        ("C2C SerDes + 连接器", "5,000", 8, "40,000", "片间互联"),
        ("VRM 供电模组", "3,000", 8, "24,000", "每 card 独立供电"),
        ("散热模组 (均热板+鳍片)", "2,000", 8, "16,000", "被动散热"),
        ("服务器平台 (CPU+BMC+DRAM+SSD)", "—", 1, "150,000", "Intel/AMD EPYC + 512GB"),
        ("机箱 + 电源 (3+1冗余) + 风扇", "—", 1, "50,000", "4U, 3kW PSU"),
        ("合计", "—", "—", f"{SERVER_COST_RMB:,d}", "整机成本估算"),
    ]

    for idx, (item, unit, qty, subtotal, note) in enumerate(cost_rows):
        row = 34 + idx
        ws.cell(row=row, column=1, value=item)
        ws.cell(row=row, column=1).font = BOLD_FONT if item == "合计" else BODY_FONT
        ws.cell(row=row, column=1).alignment = LEFT
        ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2, value=unit)
        style_data_cell(ws, row, 2, font=BOLD_FONT if item == "合计" else BODY_FONT)

        ws.cell(row=row, column=3, value=qty)
        style_data_cell(ws, row, 3, font=BOLD_FONT if item == "合计" else BODY_FONT)

        ws.cell(row=row, column=4, value=subtotal)
        style_data_cell(ws, row, 4, fmt="#,##0",
                        fill=RESULT_FILL if item == "合计" else None,
                        font=BOLD_FONT)

        ws.cell(row=row, column=5, value=note)
        ws.cell(row=row, column=5).font = BODY_FONT
        ws.cell(row=row, column=5).alignment = LEFT
        ws.cell(row=row, column=5).border = thin_border


# ============================================================================
# Sheet 5: Reference Constants
# ============================================================================

def build_reference_sheet(wb):
    ws = wb.create_sheet("参考常数")

    ws.merge_cells('A1:C1')
    ws['A1'] = "模型 & 硬件参考常数"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    sections = [
        ("DeepSeek V4 Pro 模型参数", [
            ("层数", NUM_LAYERS),
            ("隐藏维度", 7168),
            ("Expert 总数", 384),
            ("Top-K Experts", 6),
            ("MLA KV latent 维度", 512),
            ("MLA RoPE 维度", 64),
            ("KV cache 每 token 每层", f"{MLA_KV_BYTES} bytes (FP8)"),
        ]),
        ("AGM 039-F 芯片参数", [
            ("DSP 数量", "12,300"),
            ("DSP 频率", "450 MHz"),
            ("DSP MAC/cycle (fp4xfp8 原生)", 2),
            ("fp4xfp8 TMACs (原生)", "11.07"),
            ("fp8xfp8 MAC/cycle (归一化保守)", 1),
            ("fp8xfp8 TMACs (归一化)", "5.54"),
            ("fp8 TFLOPS (GPU 等效)", f"{DSP_FP8_TFLOPS_PER_CHIP:.2f}"),
            ("带宽/算力比", f"{FPGA_MBW:.1f} GB/s per TFLOP"),
            ("HBM 容量", f"{HBM_GB_PER_CHIP} GB"),
            ("HBM 带宽", "920 GB/s"),
            ("SRAM (M20K+MLAB)", "32.5 MB"),
        ]),
        ("单台服务器 (8卡x4芯片=32芯片)", [
            ("Pipeline 吞吐", f"{PIPELINE_TPS:,} tok/s"),
            ("Batch-1 延迟", f"{TOKEN_LATENCY_US} us"),
            ("Batch-1 吞吐", f"{BATCH1_TPS} tok/s"),
            ("FP8 TFLOPS (归一化)", f"{FPGA_FP8_TFLOPS_PER_SRV:.0f} (vs H200 8-GPU: {H200_FP8_TFLOPS_PER_SRV:,})"),
            ("带宽/算力比", f"{FPGA_MBW:.1f} GB/s per TFLOP (H200: {H200_MBW:.1f})"),
            ("HBM 总量", f"{CHIPS_PER_SERVER * HBM_GB_PER_CHIP:,} GB (1 TB)"),
            ("HBM 可用于 KV cache", f"~{CHIPS_PER_SERVER * HBM_KV_AVAIL_GB:.0f} GB"),
            ("功耗", f"{SERVER_POWER_KW} kW"),
            ("成本 (估算)", f"RMB {SERVER_COST_RMB:,} (100万, 含32片A7)"),
        ]),
        ("经济参数", [
            ("电价", f"{RMB_PER_KWH} 元/kWh (东数西算/内蒙)"),
            ("PUE", "1.3 (风冷数据中心)"),
            ("年运维费率", "5% CAPEX"),
            ("A7 (AGM 039-F) 芯片单价", f"RMB {A7_CHIP_COST_RMB:,d}/片"),
            ("A7 芯片成本/服务器", f"RMB {A7_CHIP_TOTAL_RMB:,d} (32片)"),
            ("FPGA 云服务定价 (混合)", "3 RMB/百万 token (DS V4 Pro 促销 5.4, 标准 21.6)"),
            ("DS V4 Pro API (标准/促销)", "I=12/3 O=24/6 RMB/1M (2026/04 官方)"),
            ("H200 8-GPU 服务器", "~300 万 RMB, ~10kW, ~2,000 tok/s (DS V4 Pro 高并发)"),
            ("Ascend 950PR 8-NPU 服务器 (估)", "~130 万 RMB, ~4kW, ~1,500 tok/s (估)"),
        ]),
    ]

    row = 3
    for title, items in sections:
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = title
        ws[f'A{row}'].font = SECTION_FONT
        row += 1

        for label, value in items:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = BODY_FONT
            ws.cell(row=row, column=1).alignment = LEFT
            ws.cell(row=row, column=1).border = thin_border

            ws.cell(row=row, column=2, value=value)
            style_data_cell(ws, row, 2)

            row += 1
        row += 1

    set_col_widths(ws, [32, 24, 24])


# ============================================================================
# Main
# ============================================================================

def main():
    wb = openpyxl.Workbook()

    build_calculator_sheet(wb)
    build_scenarios_sheet(wb)
    build_batch_sheet(wb)
    build_comparison_sheet(wb)
    build_reference_sheet(wb)

    path = "docs/fpga_supernode_sizing.xlsx"
    wb.save(path)
    print(f"Wrote {path}")
    print(f"  Sheets: {wb.sheetnames}")
    print()
    print("  Usage: Open in Excel, modify yellow cells on Sheet 1 to see sizing change.")


if __name__ == "__main__":
    main()
